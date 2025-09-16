#!/usr/bin/env python3
"""
BOM Generator v1.0 - 单文件版本
将所有代码合并到一个文件中，避免模块导入问题
"""

import tkinter as tk
from tkinter import filedialog, messagebox
import os
import sys
from typing import Dict, Any, List
import pandas as pd
import json
import openpyxl
from datetime import datetime

# 内嵌颜色代码数据（从color_codes.json提取）
COLOR_CODES = {
    "白色": "01",
    "黑色": "00", 
    "灰色": "15",
    "杏色": "70",
    "红色": "50",
    "蓝色": "40",
    "绿色": "30",
    "黄色": "60",
    "棕色": "80",
    "紫色": "90"
    # 这里应该包含完整的颜色代码映射，为简化示例只列出部分
}

class BomGenerator:
    """BOM生成器类 - 单文件版本"""
    
    # 类级别常量定义
    SHEET_NAME = '明细表'
    STYLE_CODE_COL = '款式编码'
    WAVE_COL = '波段'
    CATEGORY_COL = '品类'
    DEV_COLOR_COL = '开发颜色'
    
    # BOM模板单元格位置配置
    CELL_CONFIG = {
        'timestamp': 'J2',
        'style_code': 'B3',
        'order_type': 'F3',
        'product_name_b4': 'B4',
        'wave_info': 'F4',
        'category_info': 'J4',
    }
    
    # 预设颜色块的精确位置映射配置
    PRESET_COLOR_BLOCKS = [
        {'color_cell': 'A8', 'sku_row': 6},
        {'color_cell': 'A11', 'sku_row': 9},
        {'color_cell': 'A14', 'sku_row': 12}
    ]
    
    def __init__(self, source_path: str) -> None:
        # 设置模板路径 - 需要用户手动指定
        self.template_path = None
        self.color_codes = COLOR_CODES
        
        try:
            # 读取Excel文件
            temp_df = pd.read_excel(source_path, sheet_name=self.SHEET_NAME, header=1)
            new_columns = temp_df.iloc[0].values
            self.df = temp_df.iloc[1:].copy()
            self.df.columns = new_columns
            
            # 验证必要的列
            required_columns = [self.STYLE_CODE_COL, self.WAVE_COL, 
                              self.CATEGORY_COL, self.DEV_COLOR_COL]
            missing_columns = [col for col in required_columns if col not in self.df.columns]
            
            if missing_columns:
                raise ValueError(f"Excel文件缺少必要的列: {missing_columns}")
                
        except Exception as e:
            raise ValueError(f"读取Excel文件时发生错误: {str(e)}")
    
    def find_style_info(self, style_code: str) -> Dict[str, Any]:
        matching_rows = self.df[self.df[self.STYLE_CODE_COL] == style_code]
        
        if matching_rows.empty:
            raise ValueError(f"错误：未在源文件中找到款式编码 '{style_code}'。")
        
        row = matching_rows.iloc[0]
        return {
            self.WAVE_COL: row[self.WAVE_COL],
            self.CATEGORY_COL: row[self.CATEGORY_COL],
            self.DEV_COLOR_COL: row[self.DEV_COLOR_COL]
        }
    
    def generate_skus(self, style_code: str, dev_colors_str: str, sizes: List[str]) -> List[Dict[str, Any]]:
        result_list = []
        color_names = [color.strip() for color in dev_colors_str.split('/')]
        
        for color_name in color_names:
            try:
                color_code = self.color_codes[color_name]
            except KeyError:
                raise ValueError(f"错误：在颜色代码字典中未找到颜色 '{color_name}'。")
            
            skus_dict = {}
            for size in sizes:
                sku = f"{style_code}{color_code}{size}"
                skus_dict[size] = sku
            
            result_list.append({
                'color': color_name,
                'skus': skus_dict
            })
        
        return result_list
    
    def set_template_path(self, template_path: str):
        """设置模板文件路径"""
        self.template_path = template_path
    
    def generate_bom_file(self, style_code: str, output_dir: str) -> None:
        if not self.template_path or not os.path.exists(self.template_path):
            raise FileNotFoundError("请先选择BOM模板文件")
        
        os.makedirs(output_dir, exist_ok=True)
        
        try:
            workbook = openpyxl.load_workbook(self.template_path)
            sheet = workbook.active
        except FileNotFoundError:
            raise FileNotFoundError(f"错误：BOM模板文件未找到，路径：{self.template_path}")
        
        # 获取产品基本信息
        style_info = self.find_style_info(style_code)
        product_name = f"HECO{style_info[self.WAVE_COL]}{style_info[self.CATEGORY_COL]}{style_code}"
        
        # 填充静态字段
        config = self.CELL_CONFIG
        current_time = datetime.now().strftime("%Y/%m/%d %H:%M")
        
        sheet[config['timestamp']] = current_time
        sheet[config['style_code']] = style_code
        sheet[config['order_type']] = "首单"
        sheet[config['product_name_b4']] = product_name
        sheet[config['wave_info']] = style_info[self.WAVE_COL]
        sheet[config['category_info']] = style_info[self.CATEGORY_COL]
        
        # 生成SKU列表
        dev_colors = style_info[self.DEV_COLOR_COL]
        sizes = ['S', 'M', 'L', 'XL']
        sku_list = self.generate_skus(style_code, dev_colors, sizes)
        
        # 填充颜色和SKU信息
        for i, color_info in enumerate(sku_list):
            if i >= len(self.PRESET_COLOR_BLOCKS):
                break
            
            block_config = self.PRESET_COLOR_BLOCKS[i]
            color_cell_addr = block_config['color_cell']
            sku_target_row = block_config['sku_row']
            
            # 写入颜色名称
            sheet[color_cell_addr] = color_info['color']
            
            # 写入SKU
            sheet.cell(row=sku_target_row, column=2).value = color_info['skus']['S']
            sheet.cell(row=sku_target_row, column=3).value = color_info['skus']['M']
            sheet.cell(row=sku_target_row, column=4).value = color_info['skus']['L']
            sheet.cell(row=sku_target_row, column=5).value = color_info['skus']['XL']
        
        # 保存文件
        output_file_path = os.path.join(output_dir, f"{style_code}.xlsx")
        workbook.save(output_file_path)

class Application(tk.Tk):
    """BOM表自动生成工具的主应用程序窗口"""
    
    def __init__(self):
        super().__init__()
        
        self.title("BOM表自动生成工具 v1.0")
        self.geometry("600x400")
        
        # 状态变量
        self.source_file_path = tk.StringVar()
        self.template_file_path = tk.StringVar()
        self.output_dir_path = tk.StringVar()
        self.status_text = tk.StringVar(value="准备就绪")
        
        self._create_widgets()
    
    def _create_widgets(self):
        # 主框架
        main_frame = tk.Frame(self, padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 源文件选择
        source_frame = tk.Frame(main_frame)
        source_frame.pack(fill=tk.X, pady=(0, 15))
        
        tk.Label(source_frame, text="源文件路径:", font=("Arial", 10, "bold")).pack(anchor=tk.W)
        source_entry_frame = tk.Frame(source_frame)
        source_entry_frame.pack(fill=tk.X, pady=(5, 0))
        
        tk.Entry(source_entry_frame, textvariable=self.source_file_path, 
                state="readonly", font=("Arial", 9)).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        tk.Button(source_entry_frame, text="选择...", width=8, 
                 command=self._select_source_file).pack(side=tk.RIGHT)
        
        # 模板文件选择
        template_frame = tk.Frame(main_frame)
        template_frame.pack(fill=tk.X, pady=(0, 15))
        
        tk.Label(template_frame, text="BOM模板文件:", font=("Arial", 10, "bold")).pack(anchor=tk.W)
        template_entry_frame = tk.Frame(template_frame)
        template_entry_frame.pack(fill=tk.X, pady=(5, 0))
        
        tk.Entry(template_entry_frame, textvariable=self.template_file_path, 
                state="readonly", font=("Arial", 9)).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        tk.Button(template_entry_frame, text="选择...", width=8, 
                 command=self._select_template_file).pack(side=tk.RIGHT)
        
        # 输出目录选择
        output_frame = tk.Frame(main_frame)
        output_frame.pack(fill=tk.X, pady=(0, 20))
        
        tk.Label(output_frame, text="输出文件夹路径:", font=("Arial", 10, "bold")).pack(anchor=tk.W)
        output_entry_frame = tk.Frame(output_frame)
        output_entry_frame.pack(fill=tk.X, pady=(5, 0))
        
        tk.Entry(output_entry_frame, textvariable=self.output_dir_path,
                state="readonly", font=("Arial", 9)).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        tk.Button(output_entry_frame, text="选择...", width=8,
                 command=self._select_output_dir).pack(side=tk.RIGHT)
        
        # 操作按钮
        button_frame = tk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(10, 20))
        
        self.generate_button = tk.Button(button_frame, text="开始生成", 
                                       font=("Arial", 11, "bold"),
                                       bg="#4CAF50", fg="white",
                                       height=2, command=self._start_generation)
        self.generate_button.pack(fill=tk.X)
        
        # 状态栏
        status_frame = tk.Frame(main_frame, relief=tk.SUNKEN, bd=1)
        status_frame.pack(fill=tk.X, side=tk.BOTTOM)
        
        tk.Label(status_frame, textvariable=self.status_text,
                font=("Arial", 9), anchor=tk.W, padx=5, pady=2).pack(fill=tk.X)
    
    def _select_source_file(self):
        file_path = filedialog.askopenfilename(
            title="选择源Excel文件",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if file_path:
            self.source_file_path.set(file_path)
            self.status_text.set(f"已选择源文件: {os.path.basename(file_path)}")
    
    def _select_template_file(self):
        file_path = filedialog.askopenfilename(
            title="选择BOM模板文件",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if file_path:
            self.template_file_path.set(file_path)
            self.status_text.set(f"已选择模板文件: {os.path.basename(file_path)}")
    
    def _select_output_dir(self):
        dir_path = filedialog.askdirectory(title="选择输出文件夹")
        if dir_path:
            self.output_dir_path.set(dir_path)
            self.status_text.set(f"已选择输出目录: {os.path.basename(dir_path)}")
    
    def _start_generation(self):
        # 获取路径
        source_path = self.source_file_path.get().strip()
        template_path = self.template_file_path.get().strip()
        output_path = self.output_dir_path.get().strip()
        
        # 验证输入
        if not source_path or not template_path or not output_path:
            messagebox.showerror("错误", "请选择源文件、模板文件和输出文件夹！")
            return
        
        try:
            self.status_text.set("正在处理中，请稍候...")
            self.update()
            
            self.generate_button.config(state="disabled")
            
            # 创建生成器
            generator = BomGenerator(source_path)
            generator.set_template_path(template_path)
            
            # 获取款式编码
            style_codes = generator.df[generator.STYLE_CODE_COL].dropna().unique()
            
            if len(style_codes) == 0:
                messagebox.showwarning("警告", "源文件中没有找到任何款式编码！")
                return
            
            # 生成BOM文件
            success_count = 0
            failed_items = []
            
            for i, style_code in enumerate(style_codes):
                try:
                    self.status_text.set(f"正在处理: {style_code} ({i+1}/{len(style_codes)})")
                    self.update()
                    
                    generator.generate_bom_file(style_code, output_path)
                    success_count += 1
                    
                except Exception as e:
                    failed_items.append(f"{style_code}: {str(e)}")
                    continue
            
            # 显示结果
            if success_count == len(style_codes):
                messagebox.showinfo("成功", 
                    f"处理完成！\n\n"
                    f"成功生成 {success_count} 个BOM文件。\n"
                    f"输出目录: {output_path}")
            elif success_count > 0:
                failed_msg = "\n".join(failed_items[:5])
                if len(failed_items) > 5:
                    failed_msg += f"\n... 和其他 {len(failed_items) - 5} 个错误"
                
                messagebox.showwarning("部分成功", 
                    f"处理完成（部分成功）！\n\n"
                    f"成功: {success_count} 个\n"
                    f"失败: {len(failed_items)} 个\n\n"
                    f"失败详情:\n{failed_msg}")
            else:
                failed_msg = "\n".join(failed_items[:3])
                messagebox.showerror("失败", 
                    f"处理失败！\n\n错误详情:\n{failed_msg}")
                    
        except Exception as e:
            messagebox.showerror("发生错误", f"处理失败：\n\n{str(e)}")
            
        finally:
            self.status_text.set("准备就绪")
            self.generate_button.config(state="normal")

if __name__ == "__main__":
    app = Application()
    app.mainloop()
